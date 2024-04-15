
import matplotlib.pyplot as plt
import tkinter as tk
import openpyxl
import numpy as np 
import matplotlib.colors as mcolors

class percurso:
    def __init__(self,fig=None,ax=None,canvas=None,master=None,labeltext=None,interface=False,dados=False,fun=False,limites=None):
        self.limites=limites
        self.interface=interface
        self.labeltext=labeltext
        self.run=True
        self.canvasgraficoperc=canvas
        self.fig=fig
        self.ax=ax
        self.pontos=[]
        self.coeficientespercusos=[]
        self.dados=dados
        self.fun=fun
        self.teste()
        self.percursosfun()




        if master != None:
            master.mainloop()

    def calcular_coeficientes(self,x1, y1, x2, y2):
        # Calcular o coeficiente angular (inclinação da reta)
        a = (y2 - y1) / (x2 - x1)
        b = y1 - a * x1
        return a, b,x1,x2,y1,y2

    def verificar_entre(self, limite1,numero, limite2):
        limite_inf = min(limite1, limite2)
        limite_sup = max(limite1, limite2)
        return limite_inf <= numero <= limite_sup
    

    
    def teste(self):
        try:
            self.percursos=[]
            workbookmatriz = openpyxl.load_workbook('percursos.xlsx')
            sheet = workbookmatriz.active
            origem=sheet.cell(row=2,column=2).value,sheet.cell(row=2,column=3).value #ok
            for i in range(5,100):
                        inicio=[]
                        xi=sheet.cell(row=i,column=2).value
                        yi=sheet.cell(row=i,column=4).value
                        xf=sheet.cell(row=i,column=3).value
                        yf=sheet.cell(row=i,column=5).value
                        if xi !=None:
                            inicio=([xi,xf],[yi,yf])
                            self.percursos.append(inicio)
        except:
     
                self.percursos = [
                    ([-43.1, -40.5], [-22.8, -23.9]),
                    ([-43.1, -41.2], [-22.8, -24.0]),
                    ([-43.1, -42.8], [-22.8, -24.6]),
                    ([-43.1, -44.7], [-22.8, -25.6]),
                    ([-43.1, -41.9], [-22.8, -24.1]),
                    ([-43.1, -42.9], [-22.8, -23.8]),
                    ([-43.1, -42.7], [-22.8, -25.5]),
                    ([-43.1, -42.9], [-22.8, -25.7]),
                    ([-43.1, -40.0], [-22.8, -22.6]),
                    ([-43.1, -42.7], [-22.8, -25.4])]
            

        for rota in self.percursos:
            x1,x2,y1,y2=rota[0][0],rota[0][1],rota[1][0],rota[1][1]
            a,b,iniciox,fimx,inicioy,fimy=self.calcular_coeficientes(x1,y1,x2,y2)
            self.coeficientespercusos.append([a,b,iniciox,fimx,inicioy,fimy])


    def percursosfun(self):
        nova_origem= (-43.1, -22.8)
        origem = (-43.1, -22.8)
        diferenca_x = nova_origem[0] - origem[0]
        diferenca_y = nova_origem[1] - origem[1]

        # Ajustar as coordenadas dos percursos para a nova origem
        self.percursos_ajustados = [
            ([x + diferenca_x for x in percurso[0]], [y + diferenca_y for y in percurso[1]]) 
            for percurso in self.percursos]
        if self.interface:
            self.interfacetrue()
        
    def interfacetrue(self):
        num_cores = 20
        cores = plt.cm.get_cmap('tab20')(np.linspace(0, 1, num_cores))
        self.cores_hex = [mcolors.rgb2hex(cor) for cor in cores]
        self.cid = self.fig.canvas.mpl_connect('button_press_event', self.onclick)
        for i, (x, y) in enumerate(self.percursos_ajustados, start=1):
            plt.plot(x, y,label=f'Percurso {i}',color=self.cores_hex[i])

        

        plt.title('Percursos Percorridos')
        plt.xlabel('Longitude',fontsize=8)
        plt.ylabel('Latitude',fontsize=8)
        plt.legend(loc='center right',bbox_to_anchor=(1.25,0.5))
        
        # Limitar os intervalos dos eixos
        plt.xlim([-45, -39])  # Defina limite_min_x e limite_max_x conforme necessário
        plt.ylim([-26.5, -22.5])  # Defina limite_min_y e limite_max_y conforme necessário
        plt.subplots_adjust(right=0.8)
        plt.subplots_adjust(top=0.9)



    def lançaponto(self):
        for ponto in self.pontos:
            ponto[0].remove()

        self.fig.canvas.draw()


    def verifica_quadrado(self,x,y):
        i=0
        for limite in self.limites:
            xi=limite[0][0]
            xf=limite[0][1]
            yi=limite[1][0]
            yf=limite[1][1]
            if xi <= x <=xf:
                 if yi<=y<=yf:
                      return i

            i+=1
        

    def classifica_percurso(self,xinput,yinput):
        xult=0
        yult=0
        ultimocriterio=[[100,100,100],[xult,yult],"Percurso i"]
        percurso=1
        ultimocriterio=[[100,100,100],[xult,yult],"Percurso i"]
        percurso=1
        for coeficientes in self.coeficientespercusos:
            a,b,iniciox,fimx,inicioy,fimy=coeficientes
            yfunção=a*xinput+b
            xfunção=(yinput-b)/a

            if self.verificar_entre(inicioy , yinput , fimy):
                currentx=abs(xfunção-xinput)
                currenty=abs(yfunção-yinput)
                totaldist=(currentx**2+currenty**2)**(1/2)
                if totaldist<ultimocriterio[0][0]:
                            ultimocriterio[0][2]=currenty
                            ultimocriterio[0][1]=currentx
                            ultimocriterio[0][0]=totaldist
                            ultimocriterio[1][0],ultimocriterio[1][1],ultimocriterio[2]=xinput,yinput,percurso
            percurso+=1

        if ultimocriterio[0][0]==100:
            ultimocriterio=[[100,100,100],[xult,yult],10]
            
        return ultimocriterio
        


    def onclick(self, event):
        if event.inaxes == self.ax:
            xinput=event.xdata
            yinput=event.ydata
            ultimocriterio=self.classifica_percurso(xinput,yinput)
            limite=self.verifica_quadrado(xinput,yinput)
            current_percurso=ultimocriterio[2]
            plt.title(f'''Percurso {str(current_percurso)}  -- Limite {limite}''')
            ponto=plt.scatter(xinput,yinput,color=self.cores_hex[current_percurso])
            self.pontos.append(ponto)
            plt.draw()
            self.canvasgraficoperc.draw()
            self.fig.canvas.draw()
            self.fun(event=None)


from addplot import criagrafico
import pandas as pd
dados_excel = pd.read_excel("percursos.xlsx")
points=[]
for i in dados_excel.index:
    try:
        lo_f=dados_excel.iloc[i+2,2]
        la_f=dados_excel.iloc[i+2,4]
        int(la_f)
        points.append((lo_f,la_f))
        
    except:
        pass

origem=(-43.1,-22.8)
points.append(origem)
lado=0.15


interface=tk.Tk()
labeltextper=tk.Label(interface,text="oiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
labeltextper.pack()

plt,ax1,fig,canvas,toolbar=criagrafico(master=interface).geragrafico()
plt.scatter(-43.1,-22.8,marker='o', color='pink', label='Baía de Guanabara')
plt.tight_layout()
limites=[]
tot=len(points)
i=0
for point in points:
    if i==tot-1:
         lado+=0.15
    xi = point[0] - lado 
    xf = point[0] + lado 
    yi = point[1] - lado 
    yf = point[1] + lado 
    limx=xi,xf
    limy=yi,yf
    limites.append((limx,limy))
    x = [xi, xf, xf, xi, xi]  # Coordenadas x dos vértices do quadrado
    y = [yi, yi, yf, yf, yi]  # Coordenadas y dos vértices do quadrado
    
    plt.plot(x, y)  # Desenha o quadrado
    i+=1



atualfig=fig
atualtoolbar=toolbar
df_dadosnovos=pd.read_parquet("Dados.parquet")
percurso(fig,ax1,canvas,interface,labeltextper,interface=True,dados=df_dadosnovos,fun=None,limites=limites)
current_percuso=labeltextper.cget("text")


interface.mainloop()